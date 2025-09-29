from flask import Flask, render_template, request, redirect, url_for, flash
from openpyxl import Workbook, load_workbook
from filelock import FileLock
from datetime import datetime
import os
import uuid

app = Flask(__name__)
app.secret_key = 'dev-secret-change-me'  # change this for production use

BOOKING_FILE = 'bookings.xlsx'
LOCK_FILE = BOOKING_FILE + '.lock'

# Car database
CARS = {
    '101': {'type': 'S-Presso', 'plate': 'N140-374W'},
    '102': {'type': 'S-Presso', 'plate': 'N161-304W'},
    '103': {'type': 'S-Presso', 'plate': 'N150-785W'},
    '104': {'type': 'S-Presso', 'plate': 'N131-797W'},
    '105': {'type': 'S-Presso', 'plate': 'N160-343W'},
    '201': {'type': 'Volvo',    'plate': 'Ann 8'},
}

# Excel headers
HEADERS = [
    'BookingID', 'Timestamp', 'Name', 'Surname', 'Email', 'Phone',
    'Address', 'PO Box', 'Occupation', 'PlaceOfWork', 'AdditionalDriver',
    'StartDate', 'EndDate', 'Days', 'CarType', 'CarID', 'LicensePlate'
]

# Ensure Excel file exists
def init_excel():
    if not os.path.exists(BOOKING_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Bookings'
        ws.append(HEADERS)
        wb.save(BOOKING_FILE)

# Read bookings into list of dicts
def read_bookings():
    if not os.path.exists(BOOKING_FILE):
        return []
    wb = load_workbook(BOOKING_FILE)
    ws = wb['Bookings']
    bookings = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        b = dict(zip(HEADERS, row))
        sd = b['StartDate']
        ed = b['EndDate']
        if isinstance(sd, datetime):
            b['StartDate'] = sd.date()
        if isinstance(ed, datetime):
            b['EndDate'] = ed.date()
        bookings.append(b)
    return bookings

# Date overlap check (inclusive)
def dates_overlap(a_start, a_end, b_start, b_end):
    return not (a_end < b_start or a_start > b_end)

# Find first available car of requested type
def find_available_car(car_type, start_date, end_date, bookings):
    for car_id, info in CARS.items():
        if info['type'] != car_type:
            continue
        busy = False
        for b in bookings:
            if str(b.get('CarID')) != str(car_id):
                continue
            b_start = b['StartDate']
            b_end = b['EndDate']
            if dates_overlap(start_date, end_date, b_start, b_end):
                busy = True
                break
        if not busy:
            return car_id, info['plate']
    return None, None

# Save booking to Excel safely
def save_booking(booking_row):
    lock = FileLock(LOCK_FILE, timeout=10)
    with lock:
        init_excel()
        wb = load_workbook(BOOKING_FILE)
        ws = wb['Bookings']
        ws.append(booking_row)
        wb.save(BOOKING_FILE)

# Main route: form and submission
@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        # Get form data
        name = request.form.get('name', '').strip()
        surname = request.form.get('surname', '').strip()
        email = request.form.get('email', '').strip()
        phone = request.form.get('phone', '').strip()
        address = request.form.get('address', '').strip()
        pobox = request.form.get('pobox', '').strip()
        occupation = request.form.get('occupation', '').strip()
        placeofwork = request.form.get('placeofwork', '').strip()
        add_driver = request.form.get('additional_driver', '').strip()
        start_str = request.form.get('start_date', '').strip()
        end_str = request.form.get('end_date', '').strip()
        car_type = request.form.get('car_type', '').strip()

        # Validate fields
        if not (name and surname and email and start_str and end_str and car_type):
            flash('Please fill in all required fields.')
            return redirect(url_for('index'))

        try:
            start_date = datetime.strptime(start_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_str, '%Y-%m-%d').date()
        except ValueError:
            flash('Invalid date format.')
            return redirect(url_for('index'))

        if end_date < start_date:
            flash('End date cannot be before start date.')
            return redirect(url_for('index'))

        bookings = read_bookings()
        car_id, plate = find_available_car(car_type, start_date, end_date, bookings)
        if not car_id:
            flash(f'Sorry — no {car_type} is available for those dates.')
            return redirect(url_for('index'))

        days = (end_date - start_date).days + 1
        booking_id = uuid.uuid4().hex[:10]
        timestamp = datetime.now().isoformat()

        row = [
            booking_id, timestamp, name, surname, email, phone,
            address, pobox, occupation, placeofwork, add_driver,
            start_date, end_date, days, car_type, car_id, plate
        ]
        save_booking(row)

        booking = dict(zip(HEADERS, row))
        return render_template('success.html', booking=booking)

    return render_template('index.html')

# View all bookings
@app.route('/bookings')
def show_bookings():
    bookings = read_bookings()
    bookings_sorted = sorted(bookings, key=lambda b: b['StartDate'])
    return render_template('bookings.html', bookings=bookings_sorted)

if __name__ == '__main__':
    init_excel()
    app.run(debug=True)
